# =====================================================
# Калькулятор НМЦД
# Автор: Анна Черкасова (https://cherkasovaanna.ru/)
# 
# ⚠️ Использование в коммерческих целях — 
#    только с письменного разрешения автора.
#    Контакты: anna@cherkasovaanna.ru | ТГ @annac1119
# =====================================================

import sys
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QCheckBox, QGroupBox, QComboBox,
    QMessageBox, QDateEdit, QFileDialog
)
from PyQt6.QtCore import Qt, QDate, QLocale
from PyQt6.QtGui import QDoubleValidator

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
from datetime import datetime
import os

class NMCDCalculatorApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Расчет НМЦД (PyQt6)")
        self.setGeometry(100, 100, 700, 600) # размер окна

        self.prices = []
        self.calculated_data = None

        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout()

        # --- Общие данные ---
        general_group = QGroupBox("Общие данные")
        general_layout = QGridLayout()

        # Дата (НМЦД)
        general_layout.addWidget(QLabel("Дата (НМЦД):"), 0, 0)
        self.date_nmcd = QDateEdit(self)
        self.date_nmcd.setCalendarPopup(True)
        self.date_nmcd.setDate(QDate.currentDate())
        general_layout.addWidget(self.date_nmcd, 0, 1)

        # Item Name
        general_layout.addWidget(QLabel("Наименование предмета договора:"), 1, 0)
        self.item_name_input = QLineEdit(self)
        general_layout.addWidget(self.item_name_input, 1, 1, 1, 2)

        # Quantity
        general_layout.addWidget(QLabel("Количество:"), 2, 0)
        self.quantity_input = QLineEdit(self)
        self.quantity_input.setValidator(self.create_float_validator())
        general_layout.addWidget(self.quantity_input, 2, 1)

        # Unit of Measurement
        general_layout.addWidget(QLabel("Ед. измерения:"), 3, 0)
        self.unit_combo = QComboBox(self)
        self.unit_combo.addItems(["усл.ед", "шт", "литры", "кг"])
        general_layout.addWidget(self.unit_combo, 3, 1)

        general_group.setLayout(general_layout)
        main_layout.addWidget(general_group)

        # --- Данные поставщиков ---
        suppliers_group = QGroupBox("Данные поставщиков")
        suppliers_layout = QGridLayout()

        self.supplier_checkboxes = []
        self.supplier_name_inputs = []
        self.price_inputs = []
        self.price_labels = []

        for i in range(5):
            checkbox = QCheckBox(f"Поставщик {i+1}", self)
            checkbox.setChecked(False)
            checkbox.stateChanged.connect(lambda state, idx=i: self.toggle_supplier_fields(state, idx))
            self.supplier_checkboxes.append(checkbox)
            suppliers_layout.addWidget(checkbox, i, 0)

            name_input = QLineEdit(self)
            name_input.setPlaceholderText("Название поставщика")
            name_input.setEnabled(False)
            self.supplier_name_inputs.append(name_input)
            suppliers_layout.addWidget(name_input, i, 1)

            price_label = QLabel("Цена:")
            price_label.setEnabled(False)
            self.price_labels.append(price_label)
            suppliers_layout.addWidget(price_label, i, 2)

            price_input = QLineEdit(self)
            price_input.setPlaceholderText("Цена")
            price_input.setValidator(self.create_float_validator())
            price_input.setEnabled(False)
            self.price_inputs.append(price_input)
            suppliers_layout.addWidget(price_input, i, 3)

        suppliers_group.setLayout(suppliers_layout)
        main_layout.addWidget(suppliers_group)

        # --- Кнопки ---
        button_layout = QHBoxLayout()
        self.calculate_button = QPushButton("Рассчитать НМЦД", self)
        self.calculate_button.clicked.connect(self.calculate_nmcd)
        button_layout.addWidget(self.calculate_button)

        self.save_button = QPushButton("Сохранить в Excel", self)
        self.save_button.clicked.connect(self.save_to_excel)
        button_layout.addWidget(self.save_button)

        main_layout.addLayout(button_layout)

        self.setLayout(main_layout)

    def create_float_validator(self):
        validator = QDoubleValidator()
        validator.setLocale(QLocale(QLocale.Language.Russian, QLocale.Country.Russia))
        validator.setBottom(0.0)
        return validator

    def toggle_supplier_fields(self, state, index):
        enabled = bool(state)
        self.supplier_name_inputs[index].setEnabled(enabled)
        self.price_inputs[index].setEnabled(enabled)
        self.price_labels[index].setEnabled(enabled)

        if not enabled:
            self.supplier_name_inputs[index].clear()
            self.price_inputs[index].clear()

    def parse_float_with_comma(self, text):
        return float(text.replace(',', '.'))

    def calculate_nmcd(self):
        try:
            item_name = self.item_name_input.text()
            if not item_name:
                QMessageBox.warning(self, "Ошибка ввода", "Пожалуйста, введите наименование предмета договора.")
                return

            quantity_str = self.quantity_input.text()
            if not quantity_str:
                QMessageBox.warning(self, "Ошибка ввода", "Пожалуйста, введите количество.")
                return
            quantity = self.parse_float_with_comma(quantity_str)

            unit = self.unit_combo.currentText()
            nmcd_date_qdate = self.date_nmcd.date()
            nmcd_date = datetime(nmcd_date_qdate.year(), nmcd_date_qdate.month(), nmcd_date_qdate.day())

            self.prices = []
            supplier_names_active = []
            for i in range(5):
                if self.supplier_checkboxes[i].isChecked():
                    price_str = self.price_inputs[i].text()
                    supplier_name = self.supplier_name_inputs[i].text()

                    if not supplier_name:
                        QMessageBox.warning(self, "Ошибка ввода", f"Пожалуйста, введите наименование поставщика {i+1}.")
                        return
                    if not price_str:
                        QMessageBox.warning(self, "Ошибка ввода", f"Пожалуйста, введите цену для поставщика {i+1}.")
                        return
                    
                    self.prices.append(self.parse_float_with_comma(price_str))
                    supplier_names_active.append(supplier_name)

            if len(self.prices) < 2:
                if len(self.prices) == 1:
                    QMessageBox.information(self, "Примечание", "Вы ввели данные только для одного поставщика. При закупке у единственного поставщика Заказчик вправе определить цену, равную наименьшему значению, полученному при анализе рынка.")
                    nmcd_ryn = quantity * self.prices[0]
                    avg_price = self.prices[0]
                    s = 0.0
                    V = 0.0
                else:
                    QMessageBox.warning(self, "Ошибка", "Для расчета требуется не менее двух цен от поставщиков.")
                    return
            else:
                avg_price = sum(self.prices) / len(self.prices)
                sum_sq_diff = sum([(p - avg_price)**2 for p in self.prices])
                s = math.sqrt(sum_sq_diff / len(self.prices))

                V = (s / avg_price) * 100 if avg_price != 0 else 0

                nmcd_ryn = quantity * avg_price

            coeff_variation_warning = ""
            if V > 33:
                coeff_variation_warning = f"Внимание: Коэффициент вариации ({V:.2f}%) превышает 33%. Рекомендуется провести дополнительные исследования."

            result_text = f"Расчет НМЦД:\n" \
                          f"Среднее арифметическое цен: {avg_price:.2f}\n" \
                          f"Среднеквадратичное отклонение: {s:.2f}\n" \
                          f"Коэффициент вариации (V): {V:.2f}%\n" \
                          f"НМЦД (предварительная): {nmcd_ryn:.2f}\n\n" \
                          f"{coeff_variation_warning}"
            QMessageBox.information(self, "Результаты расчета", result_text)

            self.calculated_data = {
                "item_name": item_name,
                "quantity": quantity,
                "unit": unit,
                "prices": self.prices,
                "supplier_names": supplier_names_active,
                "avg_price": avg_price,
                "std_dev": s,
                "coeff_variation": V,
                "nmcd_ryn": nmcd_ryn,
                "nmcd_date": nmcd_date
            }

        except ValueError as ve:
            QMessageBox.critical(self, "Ошибка ввода", f"Пожалуйста, введите корректные числовые значения для количества и цен. Ошибка: {ve}")
        except Exception as e:
            QMessageBox.critical(self, "Произошла ошибка", f"Ошибка: {e}")

    def save_to_excel(self):
        if self.calculated_data is None:
            QMessageBox.warning(self, "Предупреждение", "Сначала выполните расчет НМЦД.")
            return

        try:
            data = self.calculated_data
            suggested_file_name = f"Обоснование_НМЦД_{data['item_name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(self,
                                                       "Сохранить файл обоснования НМЦД",
                                                       suggested_file_name,
                                                       "Excel Files (*.xlsx)")
            
            if not file_path:
                return

            workbook = Workbook()
            
            # --- Лист 1: Обоснование НМЦК ---
            sheet = workbook.active
            sheet.title = "Обоснование НМЦК"

            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            row_offset = 1

            sheet.cell(row=row_offset, column=1, value="Заказчик:").font = Font(bold=True)
            row_offset += 2

            header_text_part1 = "Обоснование начальной (максимальной) цены контракта / цены договора, заключаемого на"
            sheet.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=12)
            sheet.cell(row=row_offset, column=1, value=header_text_part1).font = Font(bold=True)
            sheet.cell(row=row_offset, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row_offset += 1

            sheet.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=12)
            sheet.cell(row=row_offset, column=1, value=data['item_name']).font = Font(bold=True)
            sheet.cell(row=row_offset, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            row_offset += 2

            legal_text = "Обоснование цены договора произведено методом сопоставимых рыночных цен (анализа рынка) с применением формул"
            sheet.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=12)
            sheet.cell(row=row_offset, column=1, value=legal_text).alignment = Alignment(wrap_text=True)
            row_offset += 2

            sheet.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=12)
            sheet.cell(row=row_offset, column=1, value="Расчет НМЦД методом сопоставимых рыночных цен (анализа рынка)").font = Font(bold=True)
            row_offset += 1

            supplier_headers = [data["supplier_names"][i] if i < len(data["supplier_names"]) else f"Поставщик {i+1}" for i in range(3)]
            
            headers_row_main = [
                "№ п/п",
                "Наименование, основные характеристики объекта закупки",
                "Количество товара, работы, услуги",
                "ед. измерения",
                supplier_headers[0],
                supplier_headers[1],
                supplier_headers[2],
                "Среднеквадратичное отклонение",
                "Коэффициент вариации (не должен превышать 33%)",
                "Среднее арифметическое знач",
                "Количество коммерческих предложений",
                "НМЦД ТРУ"
            ]
            sheet.append(headers_row_main)
            for col_idx in range(1, len(headers_row_main) + 1):
                sheet.cell(row=row_offset, column=col_idx).border = thin_border
                sheet.cell(row=row_offset, column=col_idx).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet.cell(row=row_offset, column=col_idx).font = Font(bold=True, size=9)
            
            row_offset += 1

            headers_row_nums = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
            sheet.append(headers_row_nums)
            for col_idx in range(1, len(headers_row_nums) + 1):
                sheet.cell(row=row_offset, column=col_idx).border = thin_border
                sheet.cell(row=row_offset, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=row_offset, column=col_idx).font = Font(size=9)
            
            row_offset += 1

            row_data_values = [
                1,
                data["item_name"],
                data["quantity"],
                data["unit"],
                data["prices"][0] if len(data["prices"]) > 0 else "",
                data["prices"][1] if len(data["prices"]) > 1 else "",
                data["prices"][2] if len(data["prices"]) > 2 else "",
                f"{data['std_dev']:.2f}",
                f"{data['coeff_variation']:.2f}%",
                f"{data['avg_price']:.2f}",
                len(data["prices"]),
                f"{data['nmcd_ryn']:.2f}"
            ]
            sheet.append(row_data_values)
            for col_idx in range(1, len(row_data_values) + 1):
                sheet.cell(row=row_offset, column=col_idx).border = thin_border
                sheet.cell(row=row_offset, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=row_offset, column=col_idx).font = Font(size=9)
            
            row_offset += 1

            sheet.append([])
            row_offset += 1
            total_row_values = ["ИТОГО:", "X", "Х", "", "", "", "", "Х", "", "Х", "Х", f"{data['nmcd_ryn']:.2f}"]
            sheet.append(total_row_values)
            sheet.cell(row=row_offset, column=1).font = Font(bold=True)
            for col_idx in range(1, len(total_row_values) + 1):
                sheet.cell(row=row_offset, column=col_idx).border = thin_border
                sheet.cell(row=row_offset, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=row_offset, column=col_idx).font = Font(bold=True, size=9)

            row_offset += 2

            sheet.cell(row=row_offset, column=1, value="Дата подготовки обоснования НМЦК:").font = Font(bold=True)
            sheet.cell(row=row_offset, column=4, value=data["nmcd_date"].strftime("%d.%m.%Y")).font = Font(bold=True)
            row_offset += 1
            sheet.cell(row=row_offset, column=1, value="Ф. И. О. исполнителя:").font = Font(bold=True)
            sheet.cell(row=row_offset, column=4, value="").font = Font(bold=True)
            
            sheet.column_dimensions['A'].width = 6
            sheet.column_dimensions['B'].width = 30
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 12
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 15
            sheet.column_dimensions['H'].width = 18
            sheet.column_dimensions['I'].width = 18
            sheet.column_dimensions['J'].width = 18
            sheet.column_dimensions['K'].width = 15
            sheet.column_dimensions['L'].width = 15

            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            sheet.page_setup.fitToPage = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0

            # --- Лист 2: Формулы расчета ---
            formulas_sheet = workbook.create_sheet("Формулы расчета")
            formulas_sheet.page_setup.orientation = formulas_sheet.ORIENTATION_PORTRAIT
            formulas_sheet.page_setup.fitToPage = True
            formulas_sheet.page_setup.fitToWidth = 1
            formulas_sheet.page_setup.fitToHeight = 0


            formulas_sheet.cell(row=1, column=1, value="Формулы расчета НМЦД").font = Font(bold=True, size=14)
            formulas_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            formulas_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

            current_row = 3

            # Среднее арифметическое цен
            formulas_sheet.cell(row=current_row, column=1, value="1. Среднее арифметическое цен (Цср):").font = Font(bold=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="Цср = (Ц1 + Ц2 + ... + Цn) / n").font = Font(italic=True)
            formulas_sheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="Где:").font = Font(bold=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="Цi - цена i-го коммерческого предложения").alignment = Alignment(indent=1)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="n - количество коммерческих предложений").alignment = Alignment(indent=1)
            current_row += 2

            # Среднеквадратичное отклонение
            formulas_sheet.cell(row=current_row, column=1, value="2. Среднеквадратичное отклонение (σ):").font = Font(bold=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="σ = √[ Σ(Цi - Цср)² / n ]").font = Font(italic=True)
            formulas_sheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 2

            # Коэффициент вариации (с добавлением примечания)
            formulas_sheet.cell(row=current_row, column=1, value="3. Коэффициент вариации (V):").font = Font(bold=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="V = (σ / Цср) * 100%").font = Font(italic=True)
            formulas_sheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="Примечание: Коэффициент вариации не должен превышать 33%").font = Font(size=9, italic=True)
            current_row += 2

            # Расчет НМЦД
            formulas_sheet.cell(row=current_row, column=1, value="4. Расчет начальной (максимальной) цены договора (НМЦД):").font = Font(bold=True)
            current_row += 1
            formulas_sheet.cell(row=current_row, column=1, value="НМЦД = Цср * Количество").font = Font(italic=True)
            formulas_sheet.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 2

            # Настройка ширины столбцов для листа с формулами
            formulas_sheet.column_dimensions['A'].width = 70
            formulas_sheet.column_dimensions['B'].width = 10
            formulas_sheet.column_dimensions['C'].width = 10
            formulas_sheet.column_dimensions['D'].width = 10

            # --- Сохранение файла Excel ---
            workbook.save(file_path)
            QMessageBox.information(self, "Успех", f"Данные успешно сохранены в файл: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "Ошибка сохранения", f"Не удалось сохранить файл Excel: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = NMCDCalculatorApp()
    window.show()
    sys.exit(app.exec())